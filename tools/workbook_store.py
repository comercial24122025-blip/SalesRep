from __future__ import annotations

import json
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


LATAM_REFERENCE_WORKBOOK = Path("/Users/erickmendez/Documents/Sales/latam_market_dashboard_automated.xlsx")
OPPORTUNITIES_WORKBOOK = Path("/Users/erickmendez/Documents/Sales/Opportunities-2026.xlsx")
DEFAULT_MISSING_DEAL_VALUE = 25000.0

STAGE_SEQUENCE = ["Lead", "Qualified", "Proposal", "Legal", "DD", "Integration", "Legal Approval", "Go Live", "Handover", "On Hold", "Closed Lost"]
STAGE_RANK = {stage: index for index, stage in enumerate(STAGE_SEQUENCE, start=1)}

MARKET_ALIASES = {
    "latam": "LATAM",
    "peru?": "Peru",
    "salvador": "El Salvador",
    ".com, pa, pe, mx, do": "Multi-market",
    "colombia - panama": "Multi-market",
    "peru - mexico": "Multi-market",
    "latam .com": "LATAM",
    "rep dominicana": "Dominican Republic",
    "dr": "Dominican Republic",
    "pr": "Puerto Rico",
    "asia": "Asia",
    "dubay": "Dubai",
}

TYPE_ALIASES = {
    "b2c": "B2C",
    "b2b": "B2B",
    "social": "Social",
    "retail": "Retail",
    "b2c + retail": "B2C + Retail",
    "b2b / b2c": "B2B / B2C",
    "b2c / b2b": "B2B / B2C",
}

PLATFORM_ALIASES = {
    "tbc": "TBC",
    "ngx": "NGX",
    "wizma": "Wizma",
    "bet construct": "BetConstruct",
    "betconstruct": "BetConstruct",
}

DEAL_COLUMNS = [
    ("Deal", "deal"),
    ("Type", "type"),
    ("Market", "market"),
    ("Platform", "platform"),
    ("Signing ETA", "signingEta"),
    ("Signing Year", "signingYear"),
    ("Signing Month", "signingMonth"),
    ("Deal Value (EUR)", "dealValue"),
    ("Legal Status", "legalStatus"),
    ("DD Status", "ddStatus"),
    ("Integration Status", "integrationStatus"),
    ("Go Live Status", "goLiveStatus"),
    ("New Traffic", "newTraffic"),
    ("Comments", "comments"),
    ("Action Items", "actionItems"),
    ("Source", "source"),
    ("Status Text", "statusText"),
    ("Stage", "stage"),
    ("Lead Flag", "leadFlag"),
    ("Signed Flag", "signedFlag"),
    ("DD Started Flag", "ddStartedFlag"),
    ("DD Completed Flag", "ddCompletedFlag"),
    ("Integration Started Flag", "integrationStartedFlag"),
    ("Integration Completed Flag", "integrationCompletedFlag"),
    ("Go Live Flag", "goLiveFlag"),
    ("Deal Value 2", "dealValueAlt"),
    ("ID", "id"),
    ("Client", "client"),
    ("Operator", "operator"),
    ("Group", "groupName"),
    ("KAM", "kam"),
    ("Jurisdiction", "jurisdiction"),
    ("Legal Entity", "legalEntity"),
    ("Site Status", "siteStatus"),
    ("Account Scope", "accountScope"),
    ("Segment", "segment"),
    ("Primary Contact", "primaryContact"),
    ("Decision Maker", "decisionMaker"),
    ("License Status", "licenseStatus"),
    ("Products Current", "productsCurrent"),
    ("Products Potential", "productsPotential"),
    ("Current Competitors", "currentCompetitors"),
    ("Target Priority", "targetPriority"),
    ("Strategic Fit", "strategicFit"),
    ("Revenue Potential EUR", "revenuePotentialEur"),
    ("Revenue Potential Score", "revenuePotentialScore"),
    ("Strategic Fit Score", "strategicFitScore"),
    ("Close Probability Score", "closeProbabilityScore"),
    ("License Score", "licenseScore"),
    ("Legal Complexity Score", "legalComplexityScore"),
    ("Technical Complexity Score", "technicalComplexityScore"),
    ("Commercial Urgency Score", "commercialUrgencyScore"),
    ("Opportunity Score", "opportunityScore"),
    ("Priority Class", "priorityClass"),
    ("Status", "status"),
    ("Mo", "month"),
    ("Evo", "evo"),
    ("Agreement", "agreement"),
    ("Integration", "integration"),
    ("DD", "dd"),
    ("Signed ETA", "signedEta"),
    ("Live Since", "liveSince"),
    ("Last Follow Up", "lastFollowUp"),
    ("Handover", "handover"),
    ("Brands", "brands"),
    ("Entity & Company Info", "entityInfo"),
    ("URL", "url"),
    ("Jira", "jira"),
    ("DD TKT", "ddTicket"),
    ("Skype", "skype"),
    ("Integration Email", "integrationEmail"),
    ("Company Name", "companyName"),
    ("Company Registration Number", "companyRegistrationNumber"),
    ("Company Registered Address", "companyRegisteredAddress"),
    ("Company Legal Representative", "companyLegalRepresentative"),
    ("Company License", "companyLicense"),
    ("Invoice Email", "invoiceEmail"),
    ("Support Email", "supportEmail"),
    ("Management Email", "managementEmail"),
    ("DD Contact Name", "ddContactName"),
    ("DD Contact Email", "ddContactEmail"),
    ("Legal Representative Name", "legalRepresentativeName"),
    ("Legal Representative ID", "legalRepresentativeId"),
    ("Legal Representative Address", "legalRepresentativeAddress"),
    ("Legal Representative Email", "legalRepresentativeEmail"),
    ("Client Based", "clientBased"),
    ("Other Live Suppliers", "otherLiveSuppliers"),
    ("Integration Team", "integrationTeam"),
    ("Teams Group", "teamsGroup"),
    ("Integration Request", "integrationRequest"),
    ("Legal Signoff Request", "legalSignoffRequest"),
    ("Other Info", "otherInfo"),
    ("Document Client Name", "documentClientName"),
    ("Proposal Validity Days", "proposalValidityDays"),
    ("Proposal Valid Until", "proposalValidUntil"),
    ("Proposal Request", "proposalRequest"),
    ("Negotiated Products", "negotiatedProducts"),
    ("Activation Requirements", "activationRequirements"),
    ("Pricing Base", "pricingBase"),
    ("Deduction Terms", "deductionTerms"),
    ("Commercial Terms", "commercialTerms"),
    ("Commercial Schedule", "commercialSchedule"),
    ("Negotiation Scope", "negotiationScope"),
    ("Setup Fee Status", "setupFeeStatus"),
    ("Setup Fee Amount", "setupFeeAmount"),
    ("Marketing Commitments", "marketingCommitments"),
    ("Live Games Top Position", "liveGamesTopPosition"),
    ("Slots Top Position", "slotsTopPosition"),
    ("Deductions Allowed", "deductionsAllowed"),
    ("Bonus Cap", "bonusCap"),
    ("Gaming Tax", "gamingTax"),
    ("Withholding Tax", "withholdingTax"),
    ("Advance Payment", "advancePayment"),
    ("Credit Notes", "creditNotes"),
    ("Updates", "updates"),
    ("Prospect Date", "prospectDate"),
    ("Offer Date", "offerDate"),
    ("DD Date", "ddDate"),
    ("Integration Date", "integrationDate"),
    ("Legal Approval Date", "legalApprovalDate"),
    ("Live Date", "liveDate"),
    ("Casino Name", "casinoName"),
    ("Ezugi ID", "ezugiId"),
    ("Evo Instance", "evoInstance"),
    ("Evo Skin ID", "evoSkinId"),
    ("Ezugi Skin", "ezugiSkin"),
    ("DB Column 1", "dbColumn1"),
    ("DB Column 2", "dbColumn2"),
    ("DB Column 3", "dbColumn3"),
    ("DB Column 4", "dbColumn4"),
    ("DB Column 5", "dbColumn5"),
    ("DB Column 6", "dbColumn6"),
    ("DB Column 7", "dbColumn7"),
    ("DB Column 8", "dbColumn8"),
    ("DB Column 9", "dbColumn9"),
    ("DB Column 10", "dbColumn10"),
    ("DB Column 11", "dbColumn11"),
    ("DB Column 12", "dbColumn12"),
    ("DB Column 13", "dbColumn13"),
    ("DB Column 14", "dbColumn14"),
]

MARKET_INTEL_COLUMNS = [
    ("ID", "id"),
    ("Country", "country"),
    ("Regulatory Status", "regulatoryStatus"),
    ("License Type", "licenseType"),
    ("Active Operators", "activeOperators"),
    ("Target Operators", "targetOperators"),
    ("Competitors Present", "competitorsPresent"),
    ("Current Products", "currentProducts"),
    ("Missing Products", "missingProducts"),
    ("Revenue Potential EUR", "revenuePotentialEur"),
    ("Regulatory Risk", "regulatoryRisk"),
    ("Opportunity Level", "opportunityLevel"),
    ("Strategic Notes", "strategicNotes"),
    ("Created At", "createdAt"),
    ("Updated At", "updatedAt"),
]

TASK_COLUMNS = [
    ("ID", "id"),
    ("Task Number", "taskNumber"),
    ("Title", "title"),
    ("Scope Type", "scopeType"),
    ("Status", "status"),
    ("Priority", "priority"),
    ("Due Date", "dueDate"),
    ("Owner", "owner"),
    ("Deal ID", "dealId"),
    ("Target ID", "targetId"),
    ("Deal", "deal"),
    ("Client", "client"),
    ("Operator", "operator"),
    ("Market", "market"),
    ("Target Year", "targetYear"),
    ("Jira Ticket", "jiraTicket"),
    ("Trace Log", "traceLog"),
    ("Next Step", "nextStep"),
    ("Notes", "notes"),
    ("Created At", "createdAt"),
    ("Updated At", "updatedAt"),
]

USER_COLUMNS = [
    ("ID", "id"),
    ("Full Name", "fullName"),
    ("Email", "email"),
    ("Role", "role"),
    ("Status", "status"),
    ("Team", "team"),
    ("Market Focus", "marketFocus"),
    ("Created At", "createdAt"),
    ("Updated At", "updatedAt"),
]

WORKSPACE_COLUMNS = [
    ("Workspace Name", "workspaceName"),
    ("Organization Name", "organizationName"),
    ("Admin Name", "adminName"),
    ("Admin Email", "adminEmail"),
    ("Subscription Plan", "subscriptionPlan"),
    ("CRM Model", "crmModel"),
    ("Fiscal Year", "fiscalYear"),
    ("Default Currency", "defaultCurrency"),
    ("Task Sequence", "taskSequence"),
]

CAMPAIGN_COLUMNS = [
    ("ID", "id"),
    ("Title", "title"),
    ("Campaign Type", "campaignType"),
    ("Status", "status"),
    ("Priority", "priority"),
    ("Operator", "operator"),
    ("Client", "client"),
    ("Deal ID", "dealId"),
    ("Deal", "deal"),
    ("Market", "market"),
    ("Owner", "owner"),
    ("Channel", "channel"),
    ("Start Date", "startDate"),
    ("End Date", "endDate"),
    ("Budget EUR", "budgetEur"),
    ("Prize Value EUR", "prizeValueEur"),
    ("Forecast Lift EUR", "forecastLiftEur"),
    ("Target Players", "targetPlayers"),
    ("Target Wager", "targetWager"),
    ("Target GGR", "targetGgr"),
    ("Jira Ticket", "jiraTicket"),
    ("Landing URL", "landingUrl"),
    ("Mechanic", "mechanic"),
    ("Offer Details", "offerDetails"),
    ("Success Metric", "successMetric"),
    ("Next Step", "nextStep"),
    ("Notes", "notes"),
    ("Trace Log", "traceLog"),
    ("Created At", "createdAt"),
    ("Updated At", "updatedAt"),
]

TARGET_METRICS = [
    ("New Signed", "newSigned", "newSignedValue"),
    ("Integrations", "integrations", "integrationsValue"),
    ("DD Pipeline", "ddPipeline", "ddPipelineValue"),
    ("New Go Live", "newGoLive", "newGoLiveValue"),
    ("Total Go Live", "totalGoLive", "totalGoLiveValue"),
]

TARGET_DETAIL_COLUMNS = [
    ("Year", "year"),
    ("Market", "market"),
    ("Type", "type"),
    ("Platform", "platform"),
    ("New Traffic", "newTraffic"),
    ("New Signed", "newSigned"),
    ("New Signed €", "newSignedValue"),
    ("Integrations", "integrations"),
    ("Integrations €", "integrationsValue"),
    ("DD Pipeline", "ddPipeline"),
    ("DD Pipeline €", "ddPipelineValue"),
    ("New Go Live", "newGoLive"),
    ("New Go Live €", "newGoLiveValue"),
    ("Total Go Live", "totalGoLive"),
    ("Total Go Live €", "totalGoLiveValue"),
]

KPI_COLUMNS = [
    ("KPI Block", "block"),
    ("KPI Name", "name"),
    ("Definition / Formula", "definition"),
    ("Stage", "stage"),
    ("Frequency", "frequency"),
    ("Notes", "notes"),
]

KPI_BOOTSTRAP = [
    {"block": "Lead", "name": "New Leads Generated", "definition": "# of new leads added in period", "stage": "Lead", "frequency": "Weekly / Monthly", "notes": "Count of distinct clients added"},
    {"block": "Qualified", "name": "Qualified Opportunities", "definition": "# of active opportunities with real commercial potential", "stage": "Qualified", "frequency": "Monthly", "notes": "Use raw status and next action"},
    {"block": "Proposal", "name": "Proposal Value (€)", "definition": "SUM of Deal Value (€) in Proposal", "stage": "Proposal", "frequency": "Monthly", "notes": "From pipeline data"},
    {"block": "Legal", "name": "Deals in Legal", "definition": "# of deals currently in Legal", "stage": "Legal", "frequency": "Weekly", "notes": "Stage = Legal"},
    {"block": "Legal", "name": "Avg Time in Legal", "definition": "Average days from first legal ticket to stage exit", "stage": "Legal", "frequency": "Weekly", "notes": "Requires start/end dates if later added"},
    {"block": "DD", "name": "Deals in DD", "definition": "# of deals currently in DD", "stage": "DD", "frequency": "Weekly", "notes": "Stage = DD"},
    {"block": "DD", "name": "DD Completion Rate %", "definition": "Deals moved from DD to Integration / Deals in DD", "stage": "DD", "frequency": "Weekly", "notes": "Process KPI"},
    {"block": "DD", "name": "DD Aging >30d", "definition": "# of DD deals with last follow-up older than 30 days", "stage": "DD", "frequency": "Weekly", "notes": "Use last follow up"},
    {"block": "Integration", "name": "Clients in Integration", "definition": "# of deals currently in Integration", "stage": "Integration", "frequency": "Weekly", "notes": "Stage = Integration"},
    {"block": "Integration", "name": "Integration Pipeline Value (€)", "definition": "SUM of Deal Value (€) in Integration", "stage": "Integration", "frequency": "Weekly", "notes": "From pipeline data"},
    {"block": "Integration", "name": "Integration to Go Live Conversion %", "definition": "Deals moved to Go Live / Deals in Integration", "stage": "Integration", "frequency": "Weekly", "notes": "Requires stage progression tracking"},
    {"block": "Go Live", "name": "New Go Lives", "definition": "# of clients with technical + legal sign-off", "stage": "Go Live", "frequency": "Monthly", "notes": "Stage = Go Live"},
    {"block": "Go Live", "name": "Go Live Value (€)", "definition": "SUM of Deal Value (€) in Go Live", "stage": "Go Live", "frequency": "Monthly", "notes": "From pipeline data"},
    {"block": "Execution", "name": "Legal to Go Live Conversion %", "definition": "Go Live deals / opportunities that entered Legal", "stage": "Legal-DD-Integration-Legal Approval-Go Live", "frequency": "Monthly", "notes": "Measures execution quality after commercial agreement"},
    {"block": "Handover", "name": "Handover Completion Rate %", "definition": "Completed Handovers / Go Live Accounts", "stage": "Handover", "frequency": "Monthly", "notes": "Measures transfer discipline after launch"},
    {"block": "Execution", "name": "Stage Cadence KPI", "definition": "Average days between recorded funnel milestone dates", "stage": "Lead-Proposal-Legal-Integration-Go Live-Handover", "frequency": "Weekly / Monthly", "notes": "Uses Prospect, Offer, Signed ETA, Integration, Live, and Handover dates; benchmark <= 30 days"},
    {"block": "Market Growth", "name": "Market Penetration %", "definition": "Live clients in market / total target operators", "stage": "Live / Growth", "frequency": "Monthly", "notes": "Target universe needed"},
    {"block": "Collaboration", "name": "Existing Client Growth (€)", "definition": "Revenue uplift influenced in existing clients", "stage": "Live / Growth", "frequency": "Monthly", "notes": "Manual input or linked rev data"},
    {"block": "Growth", "name": "Campaign Growth %", "definition": "Forecast Lift EUR / Budget EUR x 100", "stage": "Campaigns / Growth", "frequency": "Weekly / Monthly", "notes": "Projected incremental lift relative to committed campaign investment"},
]


def main() -> None:
    if len(sys.argv) < 2:
        raise SystemExit("Usage: workbook_store.py <read|write|seed-reference|import-file> ...")

    command = sys.argv[1]

    if command == "read":
        payload = read_workbook(Path(sys.argv[2]))
        print(json.dumps(payload))
        return

    if command == "write":
        workbook_path = Path(sys.argv[2])
        payload_path = Path(sys.argv[3])
        state = json.loads(payload_path.read_text())
        result = write_workbook(workbook_path, state)
        print(json.dumps(result))
        return

    if command == "seed-reference":
        print(json.dumps(load_reference_seed_state()))
        return

    if command == "import-file":
        workbook_path = Path(sys.argv[2])
        print(json.dumps(import_workbook(workbook_path)))
        return

    raise SystemExit(f"Unknown command: {command}")


def read_workbook(path: Path) -> dict[str, Any]:
    if not path.exists():
        return load_reference_seed_state()

    wb = load_workbook(path, data_only=True)
    bootstrap = create_bootstrap_state()
    deals_sheet_name = find_sheet_name(wb.sheetnames, exact={"Deals"})
    market_intel_sheet_name = find_sheet_name(wb.sheetnames, exact={"Market Intelligence"}, contains={"market intelligence"})
    targets_sheet_name = find_sheet_name(wb.sheetnames, exact={"Targets"}, contains={"targets"})
    kpi_sheet_name = find_sheet_name(wb.sheetnames, exact={"KPI Catalogue"}, contains={"kpi catalogue", "kpi catalog"})
    tasks_sheet_name = find_sheet_name(wb.sheetnames, exact={"Tasks"})
    users_sheet_name = find_sheet_name(wb.sheetnames, exact={"Users"})
    workspace_sheet_name = find_sheet_name(wb.sheetnames, exact={"Workspace"})
    campaigns_sheet_name = find_sheet_name(wb.sheetnames, exact={"Campaigns"})

    deals = normalize_deal_collection(read_deals_sheet(wb[deals_sheet_name])) if deals_sheet_name else []
    market_intel = normalize_market_intel_collection(read_market_intel_sheet(wb[market_intel_sheet_name])) if market_intel_sheet_name else normalize_market_intel_collection(bootstrap.get("marketIntel", []))
    targets = normalize_target_collection(read_targets_sheet(wb[targets_sheet_name])) if targets_sheet_name else [empty_target()]
    kpis = read_kpi_sheet(wb[kpi_sheet_name]) if kpi_sheet_name else merge_kpi_catalogue(KPI_BOOTSTRAP)
    tasks = normalize_task_collection(read_tasks_sheet(wb[tasks_sheet_name])) if tasks_sheet_name else []
    users = normalize_user_collection(read_users_sheet(wb[users_sheet_name])) if users_sheet_name else []
    if not users:
        users = normalize_user_collection(bootstrap["users"])
    workspace = normalize_workspace(read_workspace_sheet(wb[workspace_sheet_name])) if workspace_sheet_name else normalize_workspace(bootstrap["workspace"])
    campaigns = normalize_campaign_collection(read_campaigns_sheet(wb[campaigns_sheet_name])) if campaigns_sheet_name else []
    tasks, workspace = sync_task_numbers(tasks, workspace)
    return {
        "deals": deals,
        "marketIntel": market_intel,
        "targets": targets,
        "kpis": kpis,
        "tasks": tasks,
        "campaigns": campaigns,
        "users": users,
        "workspace": workspace,
        "latamReference": load_latam_reference(),
    }


def load_reference_seed_state() -> dict[str, Any]:
    bootstrap = create_bootstrap_state()
    deals = load_opportunities_deals()
    targets = load_opportunities_targets()
    latam_reference = load_latam_reference()

    return {
        "deals": normalize_deal_collection(deals or bootstrap["deals"]),
        "marketIntel": normalize_market_intel_collection(build_market_intelligence_from_reference(latam_reference) or bootstrap.get("marketIntel", [])),
        "targets": normalize_target_collection(targets or bootstrap["targets"]),
        "kpis": KPI_BOOTSTRAP,
        "tasks": normalize_task_collection(bootstrap.get("tasks", [])),
        "campaigns": normalize_campaign_collection(bootstrap.get("campaigns", [])),
        "users": normalize_user_collection(bootstrap.get("users", [])),
        "workspace": normalize_workspace(bootstrap.get("workspace", {})),
        "latamReference": latam_reference,
    }


def import_workbook(path: Path) -> dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")

    try:
        wb = load_workbook(path, data_only=True)
    except Exception as exc:
        raise ValueError(
            "SalesRep could not read this Excel file. Use a modern .xlsx/.xlsm workbook and verify the file is not corrupted or password-protected."
        ) from exc
    sheetnames = list(wb.sheetnames)
    bootstrap = create_bootstrap_state()

    salesrep_detected = any(
        [
            find_sheet_name(sheetnames, exact={"Deals"}),
            find_sheet_name(sheetnames, exact={"Tasks"}),
            find_sheet_name(sheetnames, exact={"Workspace"}),
            find_sheet_name(sheetnames, exact={"Campaigns"}),
            find_sheet_name(sheetnames, exact={"Users"}),
            find_sheet_name(sheetnames, exact={"KPI Catalogue"}, contains={"kpi catalogue", "kpi catalog"}),
            find_sheet_name(sheetnames, exact={"Market Intelligence"}, contains={"market intelligence"}),
        ]
    )
    accounts_sheet_name = find_sheet_name(sheetnames, exact={"Accounts"}, contains={"accounts"})
    hot_leads_sheet_name = find_sheet_name(sheetnames, exact={"Hot_Leads", "Hot Leads"}, contains={"hot leads"})
    targets_legacy_sheet_name = find_sheet_name(sheetnames, exact={"Targets 2025"}, contains={"targets 2025", "targets"})

    if salesrep_detected:
        state = read_workbook(path)
        return {
            "state": state,
            "meta": {
                "mode": "salesrep-workbook",
                "sheetNames": sheetnames,
                "counts": {
                    "deals": len(state.get("deals", [])),
                    "marketIntel": len(state.get("marketIntel", [])),
                    "targets": len(state.get("targets", [])),
                    "tasks": len(state.get("tasks", [])),
                    "campaigns": len(state.get("campaigns", [])),
                },
            },
        }

    if accounts_sheet_name or hot_leads_sheet_name or targets_legacy_sheet_name:
        deals_input: list[dict[str, Any]] = []
        if accounts_sheet_name:
            deals_input.extend(read_accounts_sheet(wb[accounts_sheet_name]))
        if hot_leads_sheet_name:
            deals_input.extend(read_hot_leads_sheet(wb[hot_leads_sheet_name]))

        targets_input = read_opportunities_targets_sheet(wb[targets_legacy_sheet_name]) if targets_legacy_sheet_name else []
        latam_reference = load_latam_reference()
        state = {
            "deals": normalize_deal_collection(deals_input),
            "marketIntel": normalize_market_intel_collection(build_market_intelligence_from_reference(latam_reference) or bootstrap.get("marketIntel", [])),
            "targets": normalize_target_collection(targets_input or bootstrap.get("targets", [])),
            "kpis": KPI_BOOTSTRAP,
            "tasks": normalize_task_collection(bootstrap.get("tasks", [])),
            "campaigns": normalize_campaign_collection(bootstrap.get("campaigns", [])),
            "users": normalize_user_collection(bootstrap.get("users", [])),
            "workspace": normalize_workspace(bootstrap.get("workspace", {})),
            "latamReference": latam_reference,
        }
        return {
            "state": state,
            "meta": {
                "mode": "opportunities-source",
                "sheetNames": sheetnames,
                "counts": {
                    "deals": len(state.get("deals", [])),
                    "marketIntel": len(state.get("marketIntel", [])),
                    "targets": len(state.get("targets", [])),
                    "tasks": len(state.get("tasks", [])),
                    "campaigns": len(state.get("campaigns", [])),
                },
            },
        }

    raise ValueError(
        "Unsupported Excel format. Expected a SalesRep workbook with sheets like Deals/Targets or an opportunities source workbook with Accounts/Hot_Leads/Targets 2025."
    )


def write_workbook(path: Path, state: dict[str, Any]) -> dict[str, Any]:
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    dashboard = wb.create_sheet("Dashboard")
    deals_sheet = wb.create_sheet("Deals")
    market_intel_sheet = wb.create_sheet("Market Intelligence")
    targets_sheet = wb.create_sheet("Targets")
    tasks_sheet = wb.create_sheet("Tasks")
    users_sheet = wb.create_sheet("Users")
    workspace_sheet = wb.create_sheet("Workspace")
    campaigns_sheet = wb.create_sheet("Campaigns")
    lists_sheet = wb.create_sheet("Lists")
    kpi_sheet = wb.create_sheet("KPI Catalogue")

    deals = normalize_deal_collection([normalize_deal(item) for item in state.get("deals", [])])
    market_intel = normalize_market_intel_collection([normalize_market_intel(item) for item in state.get("marketIntel", [])])
    targets = normalize_target_collection([normalize_target(item) for item in state.get("targets", [])])
    kpis = merge_kpi_catalogue([normalize_kpi(item) for item in state.get("kpis", [])] or KPI_BOOTSTRAP)
    tasks = normalize_task_collection([normalize_task(item) for item in state.get("tasks", [])])
    users = normalize_user_collection([normalize_user(item) for item in state.get("users", [])])
    if not users:
        users = normalize_user_collection(create_bootstrap_state()["users"])
    workspace = normalize_workspace(state.get("workspace", {}))
    campaigns = normalize_campaign_collection([normalize_campaign(item) for item in state.get("campaigns", [])])
    tasks, workspace = sync_task_numbers(tasks, workspace)

    write_dashboard_sheet(dashboard, deals, targets)
    write_deals_sheet(deals_sheet, deals)
    write_market_intel_sheet(market_intel_sheet, market_intel)
    write_targets_sheet(targets_sheet, targets)
    write_tasks_sheet(tasks_sheet, tasks)
    write_users_sheet(users_sheet, users)
    write_workspace_sheet(workspace_sheet, workspace)
    write_campaigns_sheet(campaigns_sheet, campaigns)
    write_lists_sheet(lists_sheet, deals, targets)
    write_kpi_sheet(kpi_sheet, kpis)

    wb.save(path)
    return {"workbookPath": str(path), "savedAt": datetime.utcnow().isoformat() + "Z"}


def write_dashboard_sheet(ws, deals, targets) -> None:
    active_year = get_active_target_year(targets)
    dominant_type = get_dominant_value([deal.get("type", "") for deal in deals]) or "All"
    target_summary = summarize_target_counts(targets, active_year)
    pipeline_value = sum_values([deal.get("dealValue") for deal in deals])
    signed_count = sum(1 for deal in deals if deal.get("signedFlag"))
    live_count = sum(1 for deal in deals if deal.get("stage") in {"Go Live", "Handover"} or deal.get("goLiveFlag"))
    dd_aging = sum(1 for deal in deals if deal.get("stage") == "DD" and days_since(deal.get("lastFollowUp")) > 30)

    rows = [
        [f"Leads, Opportunities & Target Dashboard ({active_year})"],
        [f"Default view: {active_year} | {dominant_type}"],
        ["Filters", None, None, "Primary KPIs"],
        ["Year", active_year, None, "KPI", "Count", "€ Value", "% Target"],
        ["Market", "All", None, "New Signed", signed_count, pipeline_value, target_summary["newSigned"]],
        ["Type", dominant_type, None, "Integrations", sum(1 for deal in deals if deal.get("stage") == "Integration"), sum_values([deal.get("dealValue") for deal in deals if deal.get("stage") == "Integration"]), target_summary["integrations"]],
        ["Platform", "All", None, "DD Pipeline", sum(1 for deal in deals if deal.get("stage") == "DD"), sum_values([deal.get("dealValue") for deal in deals if deal.get("stage") == "DD"]), target_summary["ddPipeline"]],
        ["New Traffic", "All", None, "New Go Live", sum(1 for deal in deals if deal.get("goLiveFlag")), sum_values([deal.get("dealValue") for deal in deals if deal.get("goLiveFlag")]), target_summary["newGoLive"]],
        ["DD Aging >30d", dd_aging, None, "Total Go Live", live_count, sum_values([deal.get("dealValue") for deal in deals if deal.get("stage") in {"Go Live", "Handover"} or deal.get("goLiveFlag")]), target_summary["totalGoLive"]],
    ]

    for row in rows:
        ws.append(row)


def write_deals_sheet(ws, deals) -> None:
    ws.append([label for label, _ in DEAL_COLUMNS])
    for deal in deals:
        ws.append([excel_cell_value(deal.get(key)) for _, key in DEAL_COLUMNS])


def write_market_intel_sheet(ws, market_intel) -> None:
    ws.append([label for label, _ in MARKET_INTEL_COLUMNS])
    for item in market_intel:
        ws.append([excel_cell_value(item.get(key)) for _, key in MARKET_INTEL_COLUMNS])


def write_targets_sheet(ws, targets) -> None:
    year = get_active_target_year(targets)
    count_summary = summarize_target_counts(targets, year)
    value_summary = summarize_target_values(targets, year)

    rows = [
        ["Target Year", year, None],
        [None, None, None],
        ["KPI", "Target Count", "Target €"],
        ["New Signed", count_summary["newSigned"], blank_if_zero(value_summary["newSignedValue"])],
        ["Integrations", count_summary["integrations"], blank_if_zero(value_summary["integrationsValue"])],
        ["DD Pipeline", count_summary["ddPipeline"], blank_if_zero(value_summary["ddPipelineValue"])],
        ["New Go Live", count_summary["newGoLive"], blank_if_zero(value_summary["newGoLiveValue"])],
        ["Total Go Live", count_summary["totalGoLive"], blank_if_zero(value_summary["totalGoLiveValue"])],
    ]

    for row in rows:
        ws.append(row)

    ws.append([None])
    ws.append([label for label, _ in TARGET_DETAIL_COLUMNS])
    for target in targets:
        ws.append([excel_cell_value(target.get(key)) for _, key in TARGET_DETAIL_COLUMNS])


def write_tasks_sheet(ws, tasks) -> None:
    ws.append([label for label, _ in TASK_COLUMNS])
    for task in tasks:
        ws.append([excel_cell_value(task.get(key)) for _, key in TASK_COLUMNS])


def write_users_sheet(ws, users) -> None:
    ws.append([label for label, _ in USER_COLUMNS])
    for user in users:
        ws.append([excel_cell_value(user.get(key)) for _, key in USER_COLUMNS])


def write_workspace_sheet(ws, workspace) -> None:
    ws.append([label for label, _ in WORKSPACE_COLUMNS])
    ws.append([excel_cell_value(workspace.get(key)) for _, key in WORKSPACE_COLUMNS])


def write_campaigns_sheet(ws, campaigns) -> None:
    ws.append([label for label, _ in CAMPAIGN_COLUMNS])
    for campaign in campaigns:
        ws.append([excel_cell_value(campaign.get(key)) for _, key in CAMPAIGN_COLUMNS])


def write_lists_sheet(ws, deals, targets) -> None:
    years = unique_values([str(target.get("year", "")) for target in targets if target.get("year")])
    markets = unique_values([deal.get("market", "") for deal in deals])
    types = unique_values([deal.get("type", "") for deal in deals])
    platforms = unique_values([deal.get("platform", "") for deal in deals])

    rows = max(len(years), len(markets), len(types), len(platforms), 1)
    for index in range(rows):
        ws.append(
            [
                "Years" if index == 0 else None,
                "All" if index == 0 else (years[index - 1] if index - 1 < len(years) else None),
                years[index] if index < len(years) else None,
                "Markets" if index == 0 else None,
                "All" if index == 0 else (markets[index - 1] if index - 1 < len(markets) else None),
                markets[index] if index < len(markets) else None,
                "Types" if index == 0 else None,
                "All" if index == 0 else (types[index - 1] if index - 1 < len(types) else None),
                types[index] if index < len(types) else None,
                "Platforms" if index == 0 else None,
                "All" if index == 0 else (platforms[index - 1] if index - 1 < len(platforms) else None),
                platforms[index] if index < len(platforms) else None,
            ]
        )


def write_kpi_sheet(ws, kpis) -> None:
    ws.append([label for label, _ in KPI_COLUMNS])
    for item in kpis:
        ws.append([item.get(key, "") for _, key in KPI_COLUMNS])


def read_deals_sheet(ws) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(cell) if cell is not None else "" for cell in rows[0]]
    index = {label: position for position, label in enumerate(header)}
    deals: list[dict[str, Any]] = []

    for row in rows[1:]:
        if not row or not any(cell is not None and str(cell).strip() for cell in row):
            continue
        raw = {key: row[index[label]] if label in index and index[label] < len(row) else None for label, key in DEAL_COLUMNS}
        deals.append(normalize_deal(raw))

    return deals


def read_market_intel_sheet(ws) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(cell) if cell is not None else "" for cell in rows[0]]
    index = {label: position for position, label in enumerate(header)}
    items: list[dict[str, Any]] = []

    for row in rows[1:]:
        if not row or not any(cell is not None and str(cell).strip() for cell in row):
            continue
        raw = {key: row[index[label]] if label in index and index[label] < len(row) else None for label, key in MARKET_INTEL_COLUMNS}
        items.append(normalize_market_intel(raw))

    return items


def read_targets_sheet(ws) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    detail_targets = read_target_detail_rows(rows)
    if detail_targets:
        return detail_targets

    year = to_int(ws["B1"].value) or datetime.utcnow().year
    summary_rows = list(ws.iter_rows(min_row=4, max_col=3, values_only=True))
    by_kpi = {stringify(row[0]).strip().lower(): row for row in summary_rows if row and stringify(row[0]).strip()}

    payload: dict[str, Any] = {
        "id": f"target-{year}-global",
        "year": year,
        "market": "Global",
        "type": "All",
        "platform": "All",
        "newTraffic": False,
    }

    for label, count_key, value_key in TARGET_METRICS:
        row = by_kpi.get(label.lower())
        if row:
            payload[count_key] = row[1]
            payload[value_key] = row[2]

    return [normalize_target(payload)]


def read_tasks_sheet(ws) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(cell) if cell is not None else "" for cell in rows[0]]
    index = {label: position for position, label in enumerate(header)}
    tasks: list[dict[str, Any]] = []

    for row in rows[1:]:
        if not row or not any(cell is not None and str(cell).strip() for cell in row):
            continue
        raw = {key: row[index[label]] if label in index and index[label] < len(row) else None for label, key in TASK_COLUMNS}
        tasks.append(normalize_task(raw))

    return tasks


def read_users_sheet(ws) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(cell) if cell is not None else "" for cell in rows[0]]
    index = {label: position for position, label in enumerate(header)}
    users: list[dict[str, Any]] = []

    for row in rows[1:]:
        if not row or not any(cell is not None and str(cell).strip() for cell in row):
            continue
        raw = {key: row[index[label]] if label in index and index[label] < len(row) else None for label, key in USER_COLUMNS}
        users.append(normalize_user(raw))

    return users


def read_workspace_sheet(ws) -> dict[str, Any]:
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return empty_workspace()

    header = [str(cell) if cell is not None else "" for cell in rows[0]]
    values = rows[1]
    index = {label: position for position, label in enumerate(header)}
    raw = {key: values[index[label]] if label in index and index[label] < len(values) else None for label, key in WORKSPACE_COLUMNS}
    return normalize_workspace(raw)


def read_campaigns_sheet(ws) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(cell) if cell is not None else "" for cell in rows[0]]
    index = {label: position for position, label in enumerate(header)}
    campaigns: list[dict[str, Any]] = []

    for row in rows[1:]:
        if not row or not any(cell is not None and str(cell).strip() for cell in row):
            continue
        raw = {key: row[index[label]] if label in index and index[label] < len(row) else None for label, key in CAMPAIGN_COLUMNS}
        campaigns.append(normalize_campaign(raw))

    return campaigns


def read_target_detail_rows(rows: list[tuple[Any, ...]]) -> list[dict[str, Any]]:
    header_index = None

    for position, row in enumerate(rows):
        header = [stringify(cell).strip() for cell in row]
        if header[: len(TARGET_DETAIL_COLUMNS)] == [label for label, _ in TARGET_DETAIL_COLUMNS]:
            header_index = position
            break

    if header_index is None:
        return []

    detail_targets: list[dict[str, Any]] = []
    labels = [label for label, _ in TARGET_DETAIL_COLUMNS]
    index = {label: position for position, label in enumerate(labels)}

    for row in rows[header_index + 1 :]:
        if not row or not any(cell not in (None, "") for cell in row):
            continue
        raw = {key: row[index[label]] if index[label] < len(row) else None for label, key in TARGET_DETAIL_COLUMNS}
        detail_targets.append(normalize_target(raw))

    return detail_targets


def read_kpi_sheet(ws) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return KPI_BOOTSTRAP

    header = [str(cell) if cell is not None else "" for cell in rows[0]]
    index = {label: position for position, label in enumerate(header)}
    items: list[dict[str, Any]] = []

    for row in rows[1:]:
        if not row or not any(cell is not None and str(cell).strip() for cell in row):
            continue
        items.append({key: stringify(row[index[label]]) if label in index and index[label] < len(row) else "" for label, key in KPI_COLUMNS})

    return merge_kpi_catalogue(items or KPI_BOOTSTRAP)


def find_sheet_name(sheetnames: list[str], *, exact: set[str] | None = None, contains: set[str] | None = None) -> str | None:
    normalized_map = {normalize_sheet_name(name): name for name in sheetnames}

    for candidate in exact or set():
        match = normalized_map.get(normalize_sheet_name(candidate))
        if match:
            return match

    contains_values = {normalize_sheet_name(value) for value in (contains or set()) if value}
    if contains_values:
        for name in sheetnames:
            normalized_name = normalize_sheet_name(name)
            if any(value in normalized_name for value in contains_values):
                return name

    return None


def normalize_sheet_name(value: Any) -> str:
    return " ".join(stringify(value).replace("_", " ").replace("-", " ").lower().split())


def load_opportunities_deals() -> list[dict[str, Any]]:
    if not OPPORTUNITIES_WORKBOOK.exists():
        return []

    wb = load_workbook(OPPORTUNITIES_WORKBOOK, data_only=True)
    accounts_sheet_name = find_sheet_name(wb.sheetnames, exact={"Accounts"}, contains={"accounts"})
    if accounts_sheet_name:
        return read_accounts_sheet(wb[accounts_sheet_name])

    return []


def load_opportunities_targets() -> list[dict[str, Any]]:
    if not OPPORTUNITIES_WORKBOOK.exists():
        return []

    wb = load_workbook(OPPORTUNITIES_WORKBOOK, data_only=True)
    targets_sheet_name = find_sheet_name(wb.sheetnames, exact={"Targets 2025"}, contains={"targets 2025", "targets"})
    if not targets_sheet_name:
        return []

    return read_opportunities_targets_sheet(wb[targets_sheet_name])


def read_opportunities_targets_sheet(ws) -> list[dict[str, Any]]:
    year = to_int(ws["C24"].value) or 2025
    targets: list[dict[str, Any]] = []

    for row_index in range(25, ws.max_row + 1):
        market = clean_text(ws.cell(row_index, 2).value)
        if not market:
            continue
        if market.upper().startswith("TARGET"):
            break

        new_signed = to_int(ws.cell(row_index, 3).value)
        if new_signed is None:
            continue

        targets.append(
            normalize_target(
                {
                    "id": f"target-{year}-{normalize_market(market).lower().replace(' ', '-')}",
                    "year": year,
                    "market": market,
                    "type": "All",
                    "platform": "All",
                    "newTraffic": False,
                    "newSigned": new_signed,
                }
            )
        )

    return targets


def read_hot_leads_sheet(ws) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [stringify(cell).strip() for cell in rows[0]]
    index = {label: position for position, label in enumerate(header) if label}
    deals: list[dict[str, Any]] = []

    for row in rows[1:]:
        client = cell_value(row, index, "Client")
        if not clean_text(client):
            continue

        market = normalize_market(cell_value(row, index, "Market"))
        if not market:
            continue

        status = clean_text(cell_value(row, index, "Status"))
        agreement = clean_text(cell_value(row, index, "Agreement"))
        integration = clean_text(cell_value(row, index, "Integration"))
        dd_value = clean_text(cell_value(row, index, "DD"))
        live_since = cell_value(row, index, "Live since")
        signed_eta = cell_value(row, index, "Signed ETA")
        last_follow_up = cell_value(row, index, "Last follow up")

        deal_value = choose_numeric_value(cell_value(row, index, "Deal_Value"), cell_value(row, index, "Deal value2"))
        stage = infer_hot_lead_stage(status, agreement, integration, dd_value, live_since)
        signed_eta_text = normalize_date(signed_eta)

        deals.append(
            normalize_deal(
                {
                    "deal": client,
                    "client": client,
                    "type": cell_value(row, index, "type"),
                    "market": market,
                    "platform": cell_value(row, index, "Platform"),
                    "stage": stage,
                    "signingEta": signed_eta_text,
                    "signedEta": signed_eta_text,
                    "signingYear": extract_year(signed_eta),
                    "signingMonth": extract_month(signed_eta),
                    "dealValue": deal_value,
                    "dealValueAlt": cell_value(row, index, "Deal value2"),
                    "status": status,
                    "agreement": agreement,
                    "integration": integration,
                    "dd": dd_value,
                    "liveSince": normalize_date(live_since),
                    "lastFollowUp": normalize_date(last_follow_up),
                    "handover": normalize_date(cell_value(row, index, "Handover")),
                    "month": cell_value(row, index, "mo"),
                    "evo": cell_value(row, index, "Evo"),
                    "brands": cell_value(row, index, "Brands"),
                    "entityInfo": cell_value(row, index, "Entity & Company Info"),
                    "url": cell_value(row, index, "URL"),
                    "jira": cell_value(row, index, "Jira"),
                    "ddTicket": cell_value(row, index, "DD TKT"),
                    "skype": cell_value(row, index, "Skype"),
                    "integrationEmail": cell_value(row, index, "Integration email"),
                    "updates": cell_value(row, index, "Updates"),
                    "source": "Opportunities-2026: Hot_Leads",
                    "statusText": build_hot_lead_status_text(status, agreement, integration, dd_value, cell_value(row, index, "Updates")),
                    "leadFlag": True,
                    "signedFlag": stage_in_or_after(stage, "Signed"),
                    "ddStartedFlag": has_process_started(dd_value) or stage_in_or_after(stage, "DD"),
                    "ddCompletedFlag": False,
                    "integrationStartedFlag": has_process_started(integration) or stage_in_or_after(stage, "Integration"),
                    "integrationCompletedFlag": stage in {"Go Live", "Live"},
                    "goLiveFlag": stage in {"Go Live", "Live"},
                    "newTraffic": clean_text(cell_value(row, index, "Evo")).lower() == "new rev",
                    "comments": status,
                    "actionItems": cell_value(row, index, "Updates"),
                }
            )
        )

    return deals


def read_accounts_sheet(ws) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [stringify(cell).strip() for cell in rows[0]]
    index = {label: position for position, label in enumerate(header) if label}
    deals: list[dict[str, Any]] = []

    for row in rows[1:]:
        site_status = clean_text(cell_value(row, index, "SITE (URL)"))
        operator_name = clean_text(cell_value(row, index, "Operator Name"))
        casino_name = clean_text(cell_value(row, index, "Casino Name (BO)"))
        deal_name = casino_name or operator_name
        if not deal_name:
            continue

        status = clean_text(cell_value(row, index, "STATUS"))
        prospect_date = cell_value(row, index, "Prospect Date")
        offer_date = cell_value(row, index, "Offer Date")
        integration_date = cell_value(row, index, "Integration Date")
        live_date = cell_value(row, index, "Live Date")
        has_progress_dates = any(normalize_date(value) for value in [offer_date, integration_date, live_date])

        if status.lower() == "" and not has_progress_dates and site_status.lower() not in {"active", "prospect", "retail", "landbased"}:
            continue
        if status.lower() in {"inactive"} and not has_progress_dates and site_status.lower() != "active":
            continue
        if status.lower() in {"canceled", "cancelled"} and not has_progress_dates and site_status.lower() != "active":
            continue

        market = normalize_market(cell_value(row, index, "Country"))
        if not market:
            continue

        stage = infer_account_stage(status, offer_date, integration_date, live_date, site_status)
        primary_date = offer_date or prospect_date
        group_name = clean_text(cell_value(row, index, "GROUP"))
        jurisdiction = clean_text(cell_value(row, index, "Jurisdiction"))
        legal_entity = clean_text(cell_value(row, index, "LEGAL ENTITY"))
        kam = clean_text(cell_value(row, index, "LATAM KAM"))
        account_scope = clean_text(cell_value(row, index, "R/.COM"))
        ezugi_id = clean_text(cell_value(row, index, "Ezugi ID"))
        evo_instance = clean_text(cell_value(row, index, "Evo Instance"))
        evo_skin_id = clean_text(cell_value(row, index, "Evo Skin ID(if applicable)"))
        ezugi_skin = clean_text(cell_value(row, index, "Ezugi Skin (if applicable)"))
        agreement_status = infer_account_agreement(status, offer_date, integration_date, live_date, site_status)
        legal_status = infer_account_legal_status(stage)
        dd_status = infer_account_dd_status(stage)
        integration_status = infer_account_integration_status(stage)
        go_live_status = infer_account_go_live_status(stage, live_date)
        last_follow_up = normalize_date(live_date or integration_date or offer_date or prospect_date)

        deals.append(
            normalize_deal(
                {
                    "deal": deal_name,
                    "client": operator_name or deal_name,
                    "operator": operator_name or deal_name,
                    "groupName": group_name,
                    "kam": kam,
                    "type": cell_value(row, index, "B2B/B2C"),
                    "market": market,
                    "platform": cell_value(row, index, "Platform"),
                    "stage": stage,
                    "signingEta": normalize_date(primary_date),
                    "signedEta": normalize_date(offer_date),
                    "signingYear": extract_year(primary_date),
                    "signingMonth": extract_month(primary_date),
                    "liveSince": normalize_date(live_date),
                    "handover": normalize_date(integration_date),
                    "lastFollowUp": last_follow_up,
                    "status": status,
                    "siteStatus": site_status,
                    "accountScope": account_scope,
                    "jurisdiction": jurisdiction,
                    "legalEntity": legal_entity,
                    "prospectDate": normalize_date(prospect_date),
                    "offerDate": normalize_date(offer_date),
                    "integrationDate": normalize_date(integration_date),
                    "liveDate": normalize_date(live_date),
                    "casinoName": casino_name,
                    "ezugiId": ezugi_id,
                    "evoInstance": evo_instance,
                    "evoSkinId": evo_skin_id,
                    "ezugiSkin": ezugi_skin,
                    "agreement": agreement_status,
                    "legalStatus": legal_status,
                    "ddStatus": dd_status,
                    "integrationStatus": integration_status,
                    "goLiveStatus": go_live_status,
                    "source": "Opportunities-2026: Accounts",
                    "statusText": build_account_status_text(site_status, status, group_name, jurisdiction, legal_entity, kam),
                    "comments": legal_entity,
                    "entityInfo": " | ".join(part for part in [jurisdiction, legal_entity] if part),
                    "brands": deal_name,
                    "url": cell_value(row, index, "URL"),
                    "dbColumn1": cell_value(row, index, "Column1"),
                    "dbColumn2": cell_value(row, index, "Column2"),
                    "dbColumn3": cell_value(row, index, "Column3"),
                    "dbColumn4": cell_value(row, index, "Column4"),
                    "dbColumn5": cell_value(row, index, "Column5"),
                    "dbColumn6": cell_value(row, index, "Column6"),
                    "dbColumn7": cell_value(row, index, "Column7"),
                    "dbColumn8": cell_value(row, index, "Column8"),
                    "dbColumn9": cell_value(row, index, "Column9"),
                    "dbColumn10": cell_value(row, index, "Column10"),
                    "dbColumn11": cell_value(row, index, "Column11"),
                    "dbColumn12": cell_value(row, index, "Column12"),
                    "dbColumn13": cell_value(row, index, "Column13"),
                    "dbColumn14": cell_value(row, index, "Column14"),
                    "leadFlag": True,
                    "signedFlag": stage_in_or_after(stage, "Signed") or agreement_status == "Signed",
                    "ddStartedFlag": stage_in_or_after(stage, "DD"),
                    "ddCompletedFlag": stage_in_or_after(stage, "Integration"),
                    "integrationStartedFlag": stage_in_or_after(stage, "Integration"),
                    "integrationCompletedFlag": stage == "Live",
                    "goLiveFlag": stage == "Live",
                    "newTraffic": account_scope.upper() == "R",
                }
            )
        )

    return deals


def load_latam_reference() -> dict[str, Any]:
    if not LATAM_REFERENCE_WORKBOOK.exists():
        return empty_latam_reference()

    wb = load_workbook(LATAM_REFERENCE_WORKBOOK, data_only=True)
    markets = read_latam_markets(wb["Chart_Data"]) if "Chart_Data" in wb.sheetnames else []
    stage_totals = read_latam_stage_totals(wb["Chart_Data"]) if "Chart_Data" in wb.sheetnames else []
    operators_by_market = read_operator_matrix(wb["Operators_by_Market"]) if "Operators_by_Market" in wb.sheetnames else []

    operator_counts = {
        entry["market"]: sum(len(entry[key]) for key in ["newBusiness", "legal", "dd", "integration", "goLive"])
        for entry in operators_by_market
    }

    enriched_markets = []
    for item in markets:
        enriched_markets.append(
            {
                **item,
                "operatorCount": operator_counts.get(item["market"], 0),
            }
        )

    return {
        "markets": enriched_markets,
        "stageTotals": stage_totals,
        "operatorsByMarket": operators_by_market,
    }


def read_latam_markets(ws) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    row_index = 2

    while row_index <= ws.max_row:
        market = stringify(ws.cell(row_index, 1).value).strip()
        if not market:
            break

        items.append(
            {
                "market": market,
                "dealCount": to_int(ws.cell(row_index, 2).value) or 0,
                "totalValue": to_float(ws.cell(row_index, 3).value) or 0,
                "stageFocus": stringify(ws.cell(row_index, 4).value),
                "priority": stringify(ws.cell(row_index, 5).value),
                "growthForecast": to_float(ws.cell(row_index, 6).value) or 0,
            }
        )
        row_index += 1

    return items


def read_latam_stage_totals(ws) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    row_index = 2

    while row_index <= ws.max_row:
        stage = stringify(ws.cell(row_index, 8).value).strip()
        if not stage:
            break

        items.append(
            {
                "stage": stage,
                "dealCount": to_int(ws.cell(row_index, 9).value) or 0,
            }
        )
        row_index += 1

    return items


def read_operator_matrix(ws) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []

    for row_index in range(2, ws.max_row + 1):
        market = stringify(ws.cell(row_index, 1).value).strip()
        if not market:
            continue

        items.append(
            {
                "market": market,
                "newBusiness": split_multiline_cell(ws.cell(row_index, 2).value),
                "legal": split_multiline_cell(ws.cell(row_index, 3).value),
                "dd": split_multiline_cell(ws.cell(row_index, 4).value),
                "integration": split_multiline_cell(ws.cell(row_index, 5).value),
                "goLive": split_multiline_cell(ws.cell(row_index, 6).value),
            }
        )

    return items


def split_multiline_cell(value: Any) -> list[str]:
    text = stringify(value).strip()
    if not text:
        return []
    return [item.strip() for item in text.splitlines() if item.strip()]


def empty_latam_reference() -> dict[str, Any]:
    return {"markets": [], "stageTotals": [], "operatorsByMarket": []}


def build_market_intelligence_from_reference(reference: dict[str, Any]) -> list[dict[str, Any]]:
    markets = reference.get("markets", []) if isinstance(reference, dict) else []
    operator_matrix = {
        clean_text(item.get("market")).lower(): item
        for item in (reference.get("operatorsByMarket", []) if isinstance(reference, dict) else [])
        if clean_text(item.get("market"))
    }

    items: list[dict[str, Any]] = []
    for market in markets:
        market_name = normalize_market(market.get("market"))
        if not market_name:
            continue

        matrix = operator_matrix.get(market_name.lower(), {})
        target_operators = unique_values(
            matrix.get("newBusiness", [])
            + matrix.get("legal", [])
            + matrix.get("dd", [])
            + matrix.get("integration", [])
            + matrix.get("goLive", [])
        )
        active_operators = unique_values(matrix.get("integration", []) + matrix.get("goLive", []))

        items.append(
            normalize_market_intel(
                {
                    "id": f"intel-{market_name.lower().replace(' ', '-')}",
                    "country": market_name,
                    "regulatoryStatus": clean_text(market.get("stageFocus")) or "Mapped",
                    "licenseType": "Market mapped",
                    "activeOperators": ", ".join(active_operators),
                    "targetOperators": ", ".join(target_operators),
                    "competitorsPresent": "",
                    "currentProducts": "",
                    "missingProducts": "",
                    "revenuePotentialEur": to_float(market.get("growthForecast")) or to_float(market.get("totalValue")) or 0,
                    "regulatoryRisk": "Medium",
                    "opportunityLevel": clean_text(market.get("priority")) or "Medium",
                    "strategicNotes": f"Stage focus: {clean_text(market.get('stageFocus')) or 'Market mapped from LATAM reference'}.",
                }
            )
        )

    return items


def create_bootstrap_state() -> dict[str, Any]:
    year = datetime.utcnow().year
    return {
        "deals": [
            normalize_deal({"id": "bootstrap-1", "deal": "Betcris.mx", "type": "B2C", "market": "Mexico", "platform": "TBC", "stage": "Integration", "signingEta": f"{year}-09-01", "signingYear": year, "signingMonth": 9, "dealValue": 55000, "legalStatus": "Signed", "integrationStatus": "Follow up integration", "source": "Signed list 1", "statusText": "Signed Follow up integration", "leadFlag": True, "signedFlag": True, "integrationStartedFlag": True}),
            normalize_deal({"id": "bootstrap-2", "deal": "Betcris.pa", "type": "B2C", "market": "Panama", "platform": "TBC", "stage": "Integration", "signingEta": f"{year}-09-01", "signingYear": year, "signingMonth": 9, "dealValue": 35000, "legalStatus": "Signed", "ddStatus": "Follow up DD", "integrationStatus": "Follow up integration", "source": "Signed list 1", "statusText": "Signed Follow up DD Follow up integration", "leadFlag": True, "signedFlag": True, "ddStartedFlag": True, "integrationStartedFlag": True}),
        ],
        "targets": [
            normalize_target(
                {
                    "id": f"target-{year}-global",
                    "year": year,
                    "market": "Global",
                    "type": "All",
                    "platform": "All",
                    "newSigned": 15,
                    "newSignedValue": 750000,
                    "integrations": 15,
                    "integrationsValue": 350000,
                    "ddPipeline": 10,
                    "ddPipelineValue": 300000,
                    "newGoLive": 15,
                    "newGoLiveValue": 500000,
                    "totalGoLive": 15,
                    "totalGoLiveValue": 750000,
                }
            )
        ],
        "kpis": KPI_BOOTSTRAP,
        "marketIntel": [
            normalize_market_intel(
                {
                    "id": "intel-mx",
                    "country": "Mexico",
                    "regulatoryStatus": "Open and active",
                    "licenseType": "B2C",
                    "activeOperators": "Betcris, Caliente, Codere",
                    "targetOperators": "Betcris, Codere, Strendus",
                    "competitorsPresent": "Pragmatic Play, Playtech",
                    "currentProducts": "Casino, sportsbook",
                    "missingProducts": "Live casino expansion, dedicated tables",
                    "revenuePotentialEur": 425000,
                    "regulatoryRisk": "Medium",
                    "opportunityLevel": "High Priority",
                    "strategicNotes": "Large market with strong operator density and visible whitespace for live product expansion.",
                }
            ),
            normalize_market_intel(
                {
                    "id": "intel-pe",
                    "country": "Peru",
                    "regulatoryStatus": "Scaling under new framework",
                    "licenseType": "Online gaming",
                    "activeOperators": "Apuesta Total, Betsson",
                    "targetOperators": "Apuesta Total, Inkabet, Betano",
                    "competitorsPresent": "Pragmatic Play, Playtech",
                    "currentProducts": "Casino, sportsbook",
                    "missingProducts": "Live casino, progressive content",
                    "revenuePotentialEur": 250000,
                    "regulatoryRisk": "Medium",
                    "opportunityLevel": "Medium",
                    "strategicNotes": "Good expansion market with clear product whitespace and strong need for structured target mapping.",
                }
            ),
        ],
        "tasks": [],
        "campaigns": [],
        "users": [
            normalize_user(
                {
                    "id": "user-admin",
                    "fullName": "LATAM Workspace Admin",
                    "email": "admin@salesrep.local",
                    "role": "Administrator",
                    "status": "Active",
                    "team": "Revenue Operations",
                    "marketFocus": "LATAM",
                }
            ),
            normalize_user(
                {
                    "id": "user-manager",
                    "fullName": "Commercial Manager",
                    "email": "manager@salesrep.local",
                    "role": "Sales Manager",
                    "status": "Active",
                    "team": "Commercial",
                    "marketFocus": "Mexico, Peru, Panama",
                }
            ),
            normalize_user(
                {
                    "id": "user-ops",
                    "fullName": "Integration Ops",
                    "email": "ops@salesrep.local",
                    "role": "Revenue Ops",
                    "status": "Active",
                    "team": "Operations",
                    "marketFocus": "Integration, Go Live",
                }
            ),
        ],
        "workspace": normalize_workspace(
            {
                "workspaceName": "SalesRep LATAM",
                "organizationName": "Evolution LATAM",
                "adminName": "LATAM Workspace Admin",
                "adminEmail": "admin@salesrep.local",
                "subscriptionPlan": "Enterprise",
                "crmModel": "New + Existing Accounts",
                "fiscalYear": year,
                "defaultCurrency": "EUR",
                "taskSequence": 0,
            }
        ),
    }


def normalize_deal(record: dict[str, Any]) -> dict[str, Any]:
    resolved_deal_value = resolve_deal_value(record)
    return {
        "id": clean_text(record.get("id")) or f"deal-{datetime.utcnow().timestamp()}",
        "deal": normalize_deal_name(record.get("deal")),
        "client": clean_text(record.get("client")),
        "type": normalize_type(record.get("type")),
        "market": normalize_market(record.get("market")),
        "platform": normalize_platform(record.get("platform")),
        "stage": normalize_stage(record.get("stage")) or "Lead",
        "signingEta": normalize_date(record.get("signingEta")),
        "signingYear": to_int(record.get("signingYear")),
        "signingMonth": to_int(record.get("signingMonth")),
        "dealValue": resolved_deal_value,
        "dealValueAlt": to_float(record.get("dealValueAlt")),
        "legalStatus": clean_text(record.get("legalStatus")),
        "ddStatus": clean_text(record.get("ddStatus")),
        "integrationStatus": clean_text(record.get("integrationStatus")),
        "goLiveStatus": clean_text(record.get("goLiveStatus")),
        "newTraffic": to_bool(record.get("newTraffic")),
        "comments": clean_text(record.get("comments")),
        "actionItems": clean_text(record.get("actionItems")),
        "source": clean_text(record.get("source")),
        "statusText": clean_text(record.get("statusText")),
        "leadFlag": to_bool(record.get("leadFlag")),
        "signedFlag": to_bool(record.get("signedFlag")),
        "ddStartedFlag": to_bool(record.get("ddStartedFlag")),
        "ddCompletedFlag": to_bool(record.get("ddCompletedFlag")),
        "integrationStartedFlag": to_bool(record.get("integrationStartedFlag")),
        "integrationCompletedFlag": to_bool(record.get("integrationCompletedFlag")),
        "goLiveFlag": to_bool(record.get("goLiveFlag")),
        "month": clean_text(record.get("month")),
        "evo": clean_text(record.get("evo")),
        "operator": clean_text(record.get("operator")),
        "groupName": clean_text(record.get("groupName")),
        "kam": clean_text(record.get("kam")),
        "jurisdiction": clean_text(record.get("jurisdiction")),
        "legalEntity": clean_text(record.get("legalEntity")),
        "siteStatus": clean_text(record.get("siteStatus")),
        "accountScope": clean_text(record.get("accountScope")),
        "segment": clean_text(record.get("segment")),
        "primaryContact": clean_text(record.get("primaryContact")),
        "decisionMaker": clean_text(record.get("decisionMaker")),
        "licenseStatus": clean_text(record.get("licenseStatus")),
        "productsCurrent": clean_text(record.get("productsCurrent")),
        "productsPotential": clean_text(record.get("productsPotential")),
        "currentCompetitors": clean_text(record.get("currentCompetitors")),
        "targetPriority": clean_text(record.get("targetPriority")),
        "strategicFit": clean_text(record.get("strategicFit")),
        "revenuePotentialEur": choose_numeric_value(record.get("revenuePotentialEur"), resolved_deal_value) or 0,
        "revenuePotentialScore": to_float(record.get("revenuePotentialScore")),
        "strategicFitScore": to_float(record.get("strategicFitScore")),
        "closeProbabilityScore": to_float(record.get("closeProbabilityScore")),
        "licenseScore": to_float(record.get("licenseScore")),
        "legalComplexityScore": to_float(record.get("legalComplexityScore")),
        "technicalComplexityScore": to_float(record.get("technicalComplexityScore")),
        "commercialUrgencyScore": to_float(record.get("commercialUrgencyScore")),
        "opportunityScore": to_float(record.get("opportunityScore")),
        "priorityClass": clean_text(record.get("priorityClass")),
        "status": clean_text(record.get("status")),
        "agreement": clean_text(record.get("agreement")),
        "integration": clean_text(record.get("integration")),
        "dd": clean_text(record.get("dd")),
        "signedEta": normalize_date(record.get("signedEta")),
        "liveSince": normalize_date(record.get("liveSince")),
        "lastFollowUp": normalize_date(record.get("lastFollowUp")),
        "handover": normalize_date(record.get("handover")),
        "brands": clean_text(record.get("brands")),
        "entityInfo": clean_text(record.get("entityInfo")),
        "url": clean_text(record.get("url")),
        "jira": clean_text(record.get("jira")),
        "ddTicket": clean_text(record.get("ddTicket")),
        "skype": clean_text(record.get("skype")),
        "integrationEmail": clean_text(record.get("integrationEmail")),
        "updates": clean_text(record.get("updates")),
        "prospectDate": normalize_date(record.get("prospectDate")),
        "offerDate": normalize_date(record.get("offerDate")),
        "ddDate": normalize_date(record.get("ddDate")),
        "integrationDate": normalize_date(record.get("integrationDate")),
        "legalApprovalDate": normalize_date(record.get("legalApprovalDate")),
        "liveDate": normalize_date(record.get("liveDate")),
        "documentClientName": clean_text(record.get("documentClientName")),
        "proposalValidityDays": to_int(record.get("proposalValidityDays")),
        "proposalValidUntil": normalize_date(record.get("proposalValidUntil")),
        "proposalRequest": clean_text(record.get("proposalRequest")),
        "negotiatedProducts": clean_text(record.get("negotiatedProducts")),
        "activationRequirements": clean_text(record.get("activationRequirements")),
        "pricingBase": clean_text(record.get("pricingBase")),
        "deductionTerms": clean_text(record.get("deductionTerms")),
        "commercialTerms": clean_text(record.get("commercialTerms")),
        "commercialSchedule": clean_text(record.get("commercialSchedule")),
        "negotiationScope": clean_text(record.get("negotiationScope")),
        "setupFeeStatus": clean_text(record.get("setupFeeStatus")),
        "setupFeeAmount": clean_text(record.get("setupFeeAmount")),
        "marketingCommitments": clean_text(record.get("marketingCommitments")),
        "liveGamesTopPosition": clean_text(record.get("liveGamesTopPosition")),
        "slotsTopPosition": clean_text(record.get("slotsTopPosition")),
        "legalSignoffRequest": clean_text(record.get("legalSignoffRequest")),
        "deductionsAllowed": clean_text(record.get("deductionsAllowed")),
        "bonusCap": clean_text(record.get("bonusCap")),
        "gamingTax": clean_text(record.get("gamingTax")),
        "withholdingTax": clean_text(record.get("withholdingTax")),
        "advancePayment": clean_text(record.get("advancePayment")),
        "creditNotes": clean_text(record.get("creditNotes")),
        "casinoName": clean_text(record.get("casinoName")),
        "ezugiId": clean_text(record.get("ezugiId")),
        "evoInstance": clean_text(record.get("evoInstance")),
        "evoSkinId": clean_text(record.get("evoSkinId")),
        "ezugiSkin": clean_text(record.get("ezugiSkin")),
        "dbColumn1": clean_text(record.get("dbColumn1")),
        "dbColumn2": clean_text(record.get("dbColumn2")),
        "dbColumn3": clean_text(record.get("dbColumn3")),
        "dbColumn4": clean_text(record.get("dbColumn4")),
        "dbColumn5": clean_text(record.get("dbColumn5")),
        "dbColumn6": clean_text(record.get("dbColumn6")),
        "dbColumn7": clean_text(record.get("dbColumn7")),
        "dbColumn8": clean_text(record.get("dbColumn8")),
        "dbColumn9": clean_text(record.get("dbColumn9")),
        "dbColumn10": clean_text(record.get("dbColumn10")),
        "dbColumn11": clean_text(record.get("dbColumn11")),
        "dbColumn12": clean_text(record.get("dbColumn12")),
        "dbColumn13": clean_text(record.get("dbColumn13")),
        "dbColumn14": clean_text(record.get("dbColumn14")),
    }


def resolve_deal_value(record: dict[str, Any]) -> float | None:
    primary_value = to_float(record.get("dealValue"))
    if primary_value is not None:
        return primary_value

    secondary_value = to_float(record.get("dealValueAlt"))
    if secondary_value is not None:
        return secondary_value

    return DEFAULT_MISSING_DEAL_VALUE if should_apply_default_deal_value(record) else None


def should_apply_default_deal_value(record: dict[str, Any]) -> bool:
    has_identity = any(clean_text(record.get(key)) for key in ["deal", "client", "operator"])
    if not has_identity:
        return False

    stage = normalize_stage(record.get("stage"))
    return stage != "Closed Lost"


def normalize_target(record: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": clean_text(record.get("id")) or f"target-{datetime.utcnow().timestamp()}",
        "year": to_int(record.get("year")) or datetime.utcnow().year,
        "market": normalize_market(record.get("market")),
        "type": normalize_type(record.get("type")),
        "platform": normalize_platform(record.get("platform")),
        "newTraffic": to_bool(record.get("newTraffic")),
        "newSigned": to_int(record.get("newSigned")) or 0,
        "newSignedValue": to_float(record.get("newSignedValue")) or 0,
        "integrations": to_int(record.get("integrations")) or 0,
        "integrationsValue": to_float(record.get("integrationsValue")) or 0,
        "ddPipeline": to_int(record.get("ddPipeline")) or 0,
        "ddPipelineValue": to_float(record.get("ddPipelineValue")) or 0,
        "newGoLive": to_int(record.get("newGoLive")) or 0,
        "newGoLiveValue": to_float(record.get("newGoLiveValue")) or 0,
        "totalGoLive": to_int(record.get("totalGoLive")) or 0,
        "totalGoLiveValue": to_float(record.get("totalGoLiveValue")) or 0,
    }


def normalize_market_intel(record: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": clean_text(record.get("id")) or f"intel-{datetime.utcnow().timestamp()}",
        "country": normalize_market(record.get("country")),
        "regulatoryStatus": clean_text(record.get("regulatoryStatus")),
        "licenseType": clean_text(record.get("licenseType")),
        "activeOperators": clean_text(record.get("activeOperators")),
        "targetOperators": clean_text(record.get("targetOperators")),
        "competitorsPresent": clean_text(record.get("competitorsPresent")),
        "currentProducts": clean_text(record.get("currentProducts")),
        "missingProducts": clean_text(record.get("missingProducts")),
        "revenuePotentialEur": to_float(record.get("revenuePotentialEur")) or 0,
        "regulatoryRisk": clean_text(record.get("regulatoryRisk")) or "Medium",
        "opportunityLevel": clean_text(record.get("opportunityLevel")) or "Medium",
        "strategicNotes": clean_text(record.get("strategicNotes")),
        "createdAt": clean_text(record.get("createdAt")),
        "updatedAt": clean_text(record.get("updatedAt")),
    }


def normalize_task(record: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": clean_text(record.get("id")) or f"task-{datetime.utcnow().timestamp()}",
        "taskNumber": clean_text(record.get("taskNumber")),
        "title": clean_text(record.get("title")),
        "scopeType": clean_text(record.get("scopeType")) or "Client",
        "status": clean_text(record.get("status")) or "Open",
        "priority": clean_text(record.get("priority")) or "Medium",
        "dueDate": normalize_date(record.get("dueDate")),
        "owner": clean_text(record.get("owner")),
        "dealId": clean_text(record.get("dealId")),
        "targetId": clean_text(record.get("targetId")),
        "deal": clean_text(record.get("deal")),
        "client": clean_text(record.get("client")),
        "operator": clean_text(record.get("operator")),
        "market": normalize_market(record.get("market")),
        "targetYear": to_int(record.get("targetYear")),
        "jiraTicket": clean_text(record.get("jiraTicket")),
        "traceLog": clean_text(record.get("traceLog")),
        "nextStep": clean_text(record.get("nextStep")),
        "notes": clean_text(record.get("notes")),
        "createdAt": clean_text(record.get("createdAt")),
        "updatedAt": clean_text(record.get("updatedAt")),
    }


def normalize_user(record: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": clean_text(record.get("id")) or f"user-{datetime.utcnow().timestamp()}",
        "fullName": clean_text(record.get("fullName")),
        "email": clean_text(record.get("email")),
        "role": clean_text(record.get("role")) or "Viewer",
        "status": clean_text(record.get("status")) or "Active",
        "team": clean_text(record.get("team")),
        "marketFocus": clean_text(record.get("marketFocus")),
        "createdAt": clean_text(record.get("createdAt")),
        "updatedAt": clean_text(record.get("updatedAt")),
    }


def normalize_workspace(record: dict[str, Any]) -> dict[str, Any]:
    year = datetime.utcnow().year
    return {
        "workspaceName": clean_text(record.get("workspaceName")) or "SalesRep LATAM",
        "organizationName": clean_text(record.get("organizationName")) or "Evolution LATAM",
        "adminName": clean_text(record.get("adminName")) or "LATAM Workspace Admin",
        "adminEmail": clean_text(record.get("adminEmail")) or "admin@salesrep.local",
        "subscriptionPlan": clean_text(record.get("subscriptionPlan")) or "Enterprise",
        "crmModel": clean_text(record.get("crmModel")) or "New + Existing Accounts",
        "fiscalYear": to_int(record.get("fiscalYear")) or year,
        "defaultCurrency": clean_text(record.get("defaultCurrency")) or "EUR",
        "taskSequence": to_int(record.get("taskSequence")) or 0,
    }


def normalize_campaign(record: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": clean_text(record.get("id")) or f"campaign-{datetime.utcnow().timestamp()}",
        "title": clean_text(record.get("title")),
        "campaignType": clean_text(record.get("campaignType")) or "Activation",
        "status": clean_text(record.get("status")) or "Planned",
        "priority": clean_text(record.get("priority")) or "Medium",
        "operator": clean_text(record.get("operator")),
        "client": clean_text(record.get("client")),
        "dealId": clean_text(record.get("dealId")),
        "deal": clean_text(record.get("deal")),
        "market": normalize_market(record.get("market")),
        "owner": clean_text(record.get("owner")),
        "channel": clean_text(record.get("channel")),
        "startDate": normalize_date(record.get("startDate")),
        "endDate": normalize_date(record.get("endDate")),
        "budgetEur": to_float(record.get("budgetEur")) or 0,
        "prizeValueEur": to_float(record.get("prizeValueEur")) or 0,
        "forecastLiftEur": to_float(record.get("forecastLiftEur")) or 0,
        "targetPlayers": to_int(record.get("targetPlayers")) or 0,
        "targetWager": to_float(record.get("targetWager")) or 0,
        "targetGgr": to_float(record.get("targetGgr")) or 0,
        "jiraTicket": clean_text(record.get("jiraTicket")),
        "landingUrl": clean_text(record.get("landingUrl")),
        "mechanic": clean_text(record.get("mechanic")),
        "offerDetails": clean_text(record.get("offerDetails")),
        "successMetric": clean_text(record.get("successMetric")),
        "nextStep": clean_text(record.get("nextStep")),
        "notes": clean_text(record.get("notes")),
        "traceLog": clean_text(record.get("traceLog")),
        "createdAt": clean_text(record.get("createdAt")),
        "updatedAt": clean_text(record.get("updatedAt")),
    }


def normalize_kpi(record: dict[str, Any]) -> dict[str, Any]:
    return {
        "block": clean_text(record.get("block")),
        "name": clean_text(record.get("name")),
        "definition": clean_text(record.get("definition")),
        "stage": clean_text(record.get("stage")),
        "frequency": clean_text(record.get("frequency")),
        "notes": clean_text(record.get("notes")),
    }


def kpi_catalogue_key(record: dict[str, Any]) -> str:
    kpi = normalize_kpi(record)
    return clean_text(kpi.get("name")).lower()


def merge_kpi_catalogue(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged: dict[str, dict[str, Any]] = {}
    preferred_order: list[str] = []

    for item in KPI_BOOTSTRAP:
        normalized = normalize_kpi(item)
        key = kpi_catalogue_key(normalized)
        preferred_order.append(key)
        merged[key] = normalized

    for item in records:
        normalized = normalize_kpi(item)
        key = kpi_catalogue_key(normalized)
        if not normalized.get("name") or key in merged:
            continue
        merged[key] = normalized

    def sort_key(item: tuple[str, dict[str, Any]]) -> tuple[int, int, str]:
        key, payload = item
        if key in preferred_order:
            return (0, preferred_order.index(key), payload.get("name", ""))
        return (1, 0, payload.get("name", ""))

    return [value for _, value in sorted(merged.items(), key=sort_key)]


def excel_cell_value(value: Any) -> Any:
    if isinstance(value, bool):
        return 1 if value else 0
    return value


def stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    return str(value)


def clean_text(value: Any) -> str:
    return " ".join(stringify(value).replace("\u00a0", " ").split())


def cell_value(row: tuple[Any, ...], index: dict[str, int], label: str) -> Any:
    position = index.get(label)
    if position is None or position >= len(row):
        return None
    return row[position]


def normalize_deal_name(value: Any) -> str:
    return clean_text(value)


def normalize_market(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return ""

    normalized = MARKET_ALIASES.get(text.lower())
    if normalized:
        return normalized

    return text


def normalize_type(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return ""

    normalized = TYPE_ALIASES.get(text.lower())
    if normalized:
        return normalized

    if "/" in text:
        parts = [TYPE_ALIASES.get(part.strip().lower(), part.strip()) for part in text.split("/")]
        unique = []
        for part in parts:
            if part and part not in unique:
                unique.append(part)
        return " / ".join(sorted(unique))

    return text


def normalize_platform(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return ""

    return PLATFORM_ALIASES.get(text.lower(), text)


def normalize_stage(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return ""

    collapsed = text.lower().replace("-", " ")
    stage_aliases = {
        "mapped lead": "Lead",
        "lead": "Lead",
        "prospect": "Lead",
        "qualified lead": "Qualified",
        "qualified": "Qualified",
        "new business": "Qualified",
        "proposal": "Proposal",
        "offer": "Proposal",
        "legal": "Legal",
        "signed": "Legal",
        "contract": "Legal",
        "negotiation": "Legal",
        "dd": "DD",
        "due diligence": "DD",
        "integration": "Integration",
        "onboarding": "Integration",
        "legal approval": "Legal Approval",
        "approval": "Legal Approval",
        "signoff": "Legal Approval",
        "sign off": "Legal Approval",
        "go live": "Go Live",
        "live": "Handover",
        "handover": "Handover",
        "on hold": "On Hold",
        "closed lost": "Closed Lost",
    }
    return stage_aliases.get(collapsed, text)


def infer_hot_lead_stage(status: Any, agreement: Any, integration: Any, dd_value: Any, live_since: Any) -> str:
    status_text = clean_text(status).lower()
    agreement_text = clean_text(agreement).lower()
    integration_text = clean_text(integration).lower()
    dd_text = clean_text(dd_value).lower()

    if normalize_date(live_since) or status_text == "live":
        return "Handover"
    if status_text in {"cancelled", "canceled", "inactive"}:
        return "Closed Lost"
    if status_text in {"hold", "on hold"}:
        return "On Hold"
    if status_text == "integration" or has_process_started(integration_text):
        return "Integration"
    if status_text == "dd" or has_process_started(dd_text):
        return "DD"
    if status_text in {"legal", "nda"}:
        return "Legal"
    if status_text == "signed":
        return "Legal"
    if agreement_text == "signed":
        return "Legal"
    return "Lead"


def infer_account_stage(status: Any, offer_date: Any, integration_date: Any, live_date: Any, site_status: Any) -> str:
    status_text = clean_text(status).lower()
    site_text = clean_text(site_status).lower()

    if normalize_date(live_date) or status_text == "live":
        return "Handover"
    if normalize_date(integration_date) or status_text == "integration":
        return "Integration"
    if status_text == "legal":
        return "Legal"
    if status_text == "offer" or normalize_date(offer_date):
        return "Proposal"
    if status_text in {"on hold", "inactive"}:
        return "On Hold"
    if status_text in {"canceled", "cancelled"} and site_text != "active":
        return "Closed Lost"

    mapping = {
        "prospect": "Lead",
        "qualified": "Qualified",
        "on hold": "On Hold",
    }
    return mapping.get(status_text, "Qualified" if clean_text(status) else "Lead")


def has_process_started(value: Any) -> bool:
    text = clean_text(value).lower()
    if not text:
        return False
    if text in {"not started", "none", "no", "n/a"}:
        return False
    return True


def stage_in_or_after(stage: Any, threshold: str) -> bool:
    return STAGE_RANK.get(normalize_stage(stage), 0) >= STAGE_RANK.get(threshold, 0)


def extract_year(value: Any) -> int | None:
    normalized = normalize_date(value)
    if not normalized:
        return None
    try:
        return datetime.fromisoformat(normalized).year
    except Exception:
        return to_int(normalized[:4])


def extract_month(value: Any) -> int | None:
    normalized = normalize_date(value)
    if not normalized:
        return None
    try:
        return datetime.fromisoformat(normalized).month
    except Exception:
        return None


def build_hot_lead_status_text(status: Any, agreement: Any, integration: Any, dd_value: Any, updates: Any) -> str:
    parts = [
        clean_text(status),
        clean_text(agreement),
        clean_text(dd_value),
        clean_text(integration),
        clean_text(updates),
    ]
    return " ".join(part for part in parts if part)


def infer_account_agreement(status: Any, offer_date: Any, integration_date: Any, live_date: Any, site_status: Any) -> str:
    status_text = clean_text(status).lower()
    site_text = clean_text(site_status).lower()
    if normalize_date(live_date) or normalize_date(integration_date):
        return "Signed"
    if status_text == "legal":
        return "Negotiation"
    if status_text == "offer" or normalize_date(offer_date):
        return "Negotiation"
    if status_text in {"on hold", "inactive"}:
        return "Blocked"
    if status_text in {"canceled", "cancelled"} and site_text != "active":
        return "Blocked"
    return "Not Started"


def infer_account_legal_status(stage: Any) -> str:
    normalized = normalize_stage(stage)
    if normalized == "Legal":
        return "In Review"
    if stage_in_or_after(normalized, "DD"):
        return "Approved"
    if normalized in {"On Hold", "Closed Lost"}:
        return "Blocked"
    return "Not Started"


def infer_account_dd_status(stage: Any) -> str:
    normalized = normalize_stage(stage)
    if normalized == "DD":
        return "In Progress"
    if stage_in_or_after(normalized, "Integration"):
        return "Completed"
    if normalized in {"On Hold", "Closed Lost"}:
        return "Blocked"
    return "Not Started"


def infer_account_integration_status(stage: Any) -> str:
    normalized = normalize_stage(stage)
    if normalized == "Integration":
        return "In Progress"
    if normalized in {"Legal Approval", "Go Live", "Handover"}:
        return "Completed"
    if normalized in {"On Hold", "Closed Lost"}:
        return "Blocked"
    return "Not Started"


def infer_account_go_live_status(stage: Any, live_date: Any) -> str:
    normalized = normalize_stage(stage)
    if normalize_date(live_date) or normalized == "Handover":
        return "Live"
    if normalized in {"Legal Approval", "Go Live"}:
        return "Legal Sign-Off"
    if normalized in {"On Hold", "Closed Lost"}:
        return "Blocked"
    return "Not Started"


def build_account_status_text(site_status: Any, status: Any, group: Any, jurisdiction: Any, legal_entity: Any, kam: Any) -> str:
    parts = [
        clean_text(site_status),
        clean_text(status),
        clean_text(group),
        clean_text(jurisdiction),
        clean_text(legal_entity),
        clean_text(kam),
    ]
    return " | ".join(part for part in parts if part)


def normalize_deal_collection(deals: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged: dict[tuple[str, str, str], dict[str, Any]] = {}

    for item in deals:
        deal = normalize_deal(item)
        key = find_matching_deal_key(merged, deal) or build_deal_key(deal)
        if key in merged:
            merged[key] = merge_deals(merged[key], deal)
        else:
            merged[key] = deal

    return sorted(merged.values(), key=deal_sort_key)


def normalize_task_collection(tasks: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged: dict[str, dict[str, Any]] = {}

    for item in tasks:
        task = normalize_task(item)
        merged[task["id"]] = task

    return sorted(merged.values(), key=task_sort_key)


def normalize_user_collection(users: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged: dict[str, dict[str, Any]] = {}

    for item in users:
        user = normalize_user(item)
        merged[user["id"]] = user

    return sorted(merged.values(), key=lambda item: (item.get("role", ""), item.get("fullName", "")))


def normalize_campaign_collection(campaigns: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged: dict[str, dict[str, Any]] = {}

    for item in campaigns:
        campaign = normalize_campaign(item)
        merged[campaign["id"]] = campaign

    return sorted(merged.values(), key=campaign_sort_key)


def normalize_market_intel_collection(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged: dict[str, dict[str, Any]] = {}

    for item in items:
        market_intel = normalize_market_intel(item)
        key = clean_text(market_intel.get("country")).lower() or market_intel.get("id", "")
        if key in merged:
            merged[key] = merge_market_intel(merged[key], market_intel)
        else:
            merged[key] = market_intel

    return sorted(merged.values(), key=lambda item: (-(to_float(item.get("revenuePotentialEur")) or 0), item.get("country", "")))


def normalize_target_collection(targets: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged: dict[tuple[int, str, str, str, bool], dict[str, Any]] = {}

    for item in targets:
        target = normalize_target(item)
        key = (
            int(target.get("year") or datetime.utcnow().year),
            target.get("market", ""),
            target.get("type", ""),
            target.get("platform", ""),
            bool(target.get("newTraffic")),
        )
        if key in merged:
            merged[key] = merge_targets(merged[key], target)
        else:
            merged[key] = target

    return sorted(merged.values(), key=lambda item: (-int(item.get("year") or 0), item.get("market", ""), item.get("type", ""), item.get("platform", "")))


def sync_task_numbers(tasks: list[dict[str, Any]], workspace: dict[str, Any]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    normalized_workspace = normalize_workspace(workspace)
    max_sequence = normalized_workspace.get("taskSequence", 0)
    normalized_tasks = [normalize_task(task) for task in tasks]

    for task in normalized_tasks:
        sequence = extract_task_sequence(task.get("taskNumber"))
        if sequence > max_sequence:
            max_sequence = sequence

    for task in sorted(normalized_tasks, key=lambda item: (clean_text(item.get("createdAt")) or "9999-12-31", item.get("title", ""))):
        if clean_text(task.get("taskNumber")):
            continue
        max_sequence += 1
        task["taskNumber"] = format_task_number(max_sequence, task, normalized_workspace)

    normalized_workspace["taskSequence"] = max_sequence
    return normalize_task_collection(normalized_tasks), normalized_workspace


def extract_task_sequence(value: Any) -> int:
    text = clean_text(value)
    if not text or "-" not in text:
        return 0
    tail = text.rsplit("-", 1)[-1]
    return to_int(tail) or 0


def format_task_number(sequence: int, task: dict[str, Any], workspace: dict[str, Any]) -> str:
    year = to_int(task.get("targetYear")) or to_int(workspace.get("fiscalYear")) or datetime.utcnow().year
    return f"TSK-{year}-{sequence:04d}"


def build_deal_key(deal: dict[str, Any]) -> tuple[str, str, str]:
    market_key = clean_text(deal.get("market")).lower()
    platform_key = clean_text(deal.get("platform")).lower()
    return (
        clean_text(deal.get("deal")).lower(),
        market_key,
        "" if market_key else platform_key,
    )


def find_matching_deal_key(merged: dict[tuple[str, str, str], dict[str, Any]], deal: dict[str, Any]) -> tuple[str, str, str] | None:
    exact_key = build_deal_key(deal)
    if exact_key in merged:
        return exact_key

    deal_name, market_name, platform_name = exact_key
    for current_key in merged:
        current_deal, current_market, current_platform = current_key
        if current_deal != deal_name:
            continue
        if market_name and current_market == market_name:
            return current_key
        if not market_name and (not platform_name or not current_platform or platform_name == current_platform):
            return current_key

    return None


def deal_sort_key(deal: dict[str, Any]) -> tuple[int, str, str]:
    return (
        -STAGE_RANK.get(deal.get("stage", ""), 0),
        deal.get("market", ""),
        deal.get("deal", ""),
    )


def task_sort_key(task: dict[str, Any]) -> tuple[int, int, str, str]:
    status_order = {"Open": 0, "In Progress": 1, "Waiting": 2, "Blocked": 3, "Done": 4}
    priority_order = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}
    due_date = clean_text(task.get("dueDate")) or "9999-12-31"
    return (
        status_order.get(task.get("status", ""), 9),
        priority_order.get(task.get("priority", ""), 9),
        due_date,
        task.get("title", ""),
    )


def campaign_sort_key(campaign: dict[str, Any]) -> tuple[int, int, str, str]:
    status_order = {"Live": 0, "Ready": 1, "Planned": 2, "Paused": 3, "Completed": 4, "Cancelled": 5}
    priority_order = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}
    start_date = clean_text(campaign.get("startDate")) or "9999-12-31"
    return (
        status_order.get(campaign.get("status", ""), 9),
        priority_order.get(campaign.get("priority", ""), 9),
        start_date,
        campaign.get("title", ""),
    )


def merge_deals(left: dict[str, Any], right: dict[str, Any]) -> dict[str, Any]:
    preferred = pick_preferred_deal(left, right)
    secondary = right if preferred is left else left
    result = dict(preferred)

    for key in [
        "deal",
        "client",
        "operator",
        "groupName",
        "kam",
        "market",
        "jurisdiction",
        "legalEntity",
        "siteStatus",
        "accountScope",
        "segment",
        "primaryContact",
        "decisionMaker",
        "licenseStatus",
        "targetPriority",
        "priorityClass",
        "casinoName",
        "legalStatus",
        "ddStatus",
        "integrationStatus",
        "goLiveStatus",
        "month",
        "evo",
        "status",
        "agreement",
        "integration",
        "dd",
        "brands",
        "entityInfo",
        "url",
        "jira",
        "ddTicket",
        "skype",
        "integrationEmail",
    ]:
        result[key] = choose_preferred_string(preferred.get(key), secondary.get(key))

    result["platform"] = merge_enum_values(left.get("platform"), right.get("platform"))

    for key in [
        "source",
        "statusText",
        "comments",
        "actionItems",
        "updates",
        "legalSignoffRequest",
        "productsCurrent",
        "productsPotential",
        "currentCompetitors",
        "strategicFit",
        "ezugiId",
        "evoInstance",
        "evoSkinId",
        "ezugiSkin",
        "dbColumn1",
        "dbColumn2",
        "dbColumn3",
        "dbColumn4",
        "dbColumn5",
        "dbColumn6",
        "dbColumn7",
        "dbColumn8",
        "dbColumn9",
        "dbColumn10",
        "dbColumn11",
        "dbColumn12",
        "dbColumn13",
        "dbColumn14",
    ]:
        result[key] = merge_text_blocks(preferred.get(key), secondary.get(key))

    for key in [
        "dealValue",
        "dealValueAlt",
        "revenuePotentialEur",
        "revenuePotentialScore",
        "strategicFitScore",
        "closeProbabilityScore",
        "licenseScore",
        "legalComplexityScore",
        "technicalComplexityScore",
        "commercialUrgencyScore",
        "opportunityScore",
    ]:
        result[key] = choose_numeric_value(left.get(key), right.get(key))

    for key in ["signingYear", "signingMonth"]:
        result[key] = choose_numeric_value(left.get(key), right.get(key))

    for key in ["signingEta", "signedEta", "liveSince", "lastFollowUp", "handover", "prospectDate", "offerDate", "ddDate", "integrationDate", "legalApprovalDate", "liveDate", "proposalValidUntil"]:
        result[key] = choose_preferred_string(preferred.get(key), secondary.get(key))

    for key in ["newTraffic", "leadFlag", "signedFlag", "ddStartedFlag", "ddCompletedFlag", "integrationStartedFlag", "integrationCompletedFlag", "goLiveFlag"]:
        result[key] = bool(left.get(key) or right.get(key))

    result["stage"] = choose_stage(left.get("stage"), right.get("stage"))
    result["type"] = merge_enum_values(left.get("type"), right.get("type"))
    result["id"] = choose_preferred_string(left.get("id"), right.get("id")) or result["id"]

    return normalize_deal(result)


def merge_targets(left: dict[str, Any], right: dict[str, Any]) -> dict[str, Any]:
    result = dict(left)
    result["id"] = choose_preferred_string(left.get("id"), right.get("id")) or result["id"]

    for key in ["newSigned", "newSignedValue", "integrations", "integrationsValue", "ddPipeline", "ddPipelineValue", "newGoLive", "newGoLiveValue", "totalGoLive", "totalGoLiveValue"]:
        result[key] = (to_float(left.get(key)) or 0) + (to_float(right.get(key)) or 0)
        if key.endswith("Value"):
            continue
        result[key] = int(result[key])

    return normalize_target(result)


def merge_market_intel(left: dict[str, Any], right: dict[str, Any]) -> dict[str, Any]:
    result = dict(left)
    result["id"] = choose_preferred_string(left.get("id"), right.get("id")) or result["id"]

    for key in [
        "country",
        "regulatoryStatus",
        "licenseType",
        "activeOperators",
        "targetOperators",
        "competitorsPresent",
        "currentProducts",
        "missingProducts",
        "regulatoryRisk",
        "opportunityLevel",
        "strategicNotes",
        "createdAt",
        "updatedAt",
    ]:
        result[key] = choose_preferred_string(left.get(key), right.get(key))

    result["revenuePotentialEur"] = choose_numeric_value(left.get("revenuePotentialEur"), right.get("revenuePotentialEur")) or 0
    return normalize_market_intel(result)


def pick_preferred_deal(left: dict[str, Any], right: dict[str, Any]) -> dict[str, Any]:
    left_score = deal_completeness_score(left)
    right_score = deal_completeness_score(right)
    if right_score > left_score:
        return right
    return left


def deal_completeness_score(deal: dict[str, Any]) -> int:
    score = STAGE_RANK.get(deal.get("stage", ""), 0) * 10
    score += 8 if to_float(deal.get("dealValue")) is not None else 0
    score += 5 if deal.get("signedFlag") else 0
    score += 4 if deal.get("integrationStartedFlag") else 0
    score += 3 if deal.get("ddStartedFlag") else 0
    score += 2 if deal.get("goLiveFlag") else 0
    score += sum(
        1
        for key in [
            "statusText",
            "comments",
            "actionItems",
            "source",
            "client",
            "operator",
            "kam",
            "segment",
            "decisionMaker",
            "licenseStatus",
            "legalEntity",
            "url",
            "casinoName",
            "ezugiId",
            "evoInstance",
        ]
        if clean_text(deal.get(key))
    )
    return score


def choose_preferred_string(left: Any, right: Any) -> str:
    left_text = clean_text(left)
    right_text = clean_text(right)
    if not left_text:
        return right_text
    if not right_text:
        return left_text
    return left_text if display_text_score(left_text) >= display_text_score(right_text) else right_text


def display_text_score(text: str) -> int:
    score = len(text)
    if not text.isupper():
        score += 4
    if any(char.islower() for char in text) and any(char.isupper() for char in text):
        score += 3
    if text == clean_text(text):
        score += 1
    return score


def merge_text_blocks(left: Any, right: Any) -> str:
    values = []
    for item in [left, right]:
        text = clean_text(item)
        if text and text not in values:
            values.append(text)
    return " | ".join(values)


def choose_numeric_value(left: Any, right: Any) -> int | float | None:
    left_number = to_float(left)
    right_number = to_float(right)
    if left_number is None:
        return right_number
    if right_number is None:
        return left_number
    if right_number > left_number:
        return right_number
    return left_number


def choose_stage(left: Any, right: Any) -> str:
    left_stage = normalize_stage(left)
    right_stage = normalize_stage(right)
    if STAGE_RANK.get(right_stage, 0) > STAGE_RANK.get(left_stage, 0):
        return right_stage
    return left_stage or right_stage


def merge_enum_values(left: Any, right: Any) -> str:
    values = []
    for item in [left, right]:
        text = clean_text(item)
        if text and text not in values:
            values.append(text)
    if not values:
        return ""
    if len(values) == 1:
        return values[0]
    return " / ".join(sorted(values))


def normalize_date(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = stringify(value).strip()
    if not text:
        return ""
    if len(text) >= 10 and text[4] == "-":
        return text[:10]
    for pattern in ("%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, pattern).date().isoformat()
        except Exception:
            continue
    return text


def to_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    return stringify(value).strip().lower() in {"true", "yes", "1"}


def to_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(float(value))
    except Exception:
        return None


def to_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except Exception:
        return None


def sum_values(values: list[Any]) -> float:
    total = 0.0
    for value in values:
        number = to_float(value)
        if number is not None:
            total += number
    return total


def get_active_target_year(targets: list[dict[str, Any]]) -> int:
    years = [target.get("year") for target in targets if target.get("year")]
    return max(years) if years else datetime.utcnow().year


def summarize_target_counts(targets: list[dict[str, Any]], year: int) -> dict[str, int]:
    scoped = [target for target in targets if int(target.get("year", year)) == year]
    return {
        "newSigned": sum(int(target.get("newSigned", 0)) for target in scoped),
        "integrations": sum(int(target.get("integrations", 0)) for target in scoped),
        "ddPipeline": sum(int(target.get("ddPipeline", 0)) for target in scoped),
        "newGoLive": sum(int(target.get("newGoLive", 0)) for target in scoped),
        "totalGoLive": sum(int(target.get("totalGoLive", 0)) for target in scoped),
    }


def summarize_target_values(targets: list[dict[str, Any]], year: int) -> dict[str, float]:
    scoped = [target for target in targets if int(target.get("year", year)) == year]
    return {
        "newSignedValue": sum_values([target.get("newSignedValue") for target in scoped]),
        "integrationsValue": sum_values([target.get("integrationsValue") for target in scoped]),
        "ddPipelineValue": sum_values([target.get("ddPipelineValue") for target in scoped]),
        "newGoLiveValue": sum_values([target.get("newGoLiveValue") for target in scoped]),
        "totalGoLiveValue": sum_values([target.get("totalGoLiveValue") for target in scoped]),
    }


def unique_values(values: list[str]) -> list[str]:
    return sorted({value for value in values if value})


def get_dominant_value(values: list[str]) -> str:
    counts: dict[str, int] = {}
    for value in values:
        if value:
            counts[value] = counts.get(value, 0) + 1
    return max(counts, key=counts.get) if counts else ""


def blank_if_zero(value: Any) -> Any:
    number = to_float(value)
    if number in (None, 0):
        return None
    return number


def empty_target() -> dict[str, Any]:
    year = datetime.utcnow().year
    return normalize_target({"id": f"target-{year}-global", "year": year, "market": "Global", "type": "All", "platform": "All"})


def empty_workspace() -> dict[str, Any]:
    return normalize_workspace({})


def targets_have_values(targets: list[dict[str, Any]]) -> bool:
    for target in targets:
        for _, count_key, value_key in TARGET_METRICS:
            if int(target.get(count_key, 0) or 0) > 0:
                return True
            if (to_float(target.get(value_key)) or 0) > 0:
                return True
    return False


def days_since(date_text: Any) -> int:
    text = normalize_date(date_text)
    if not text:
        return 0
    try:
        current = datetime.utcnow().date()
        target = datetime.fromisoformat(text).date()
        return (current - target).days
    except Exception:
        return 0


if __name__ == "__main__":
    main()
